import subprocess
import sys
import calendar
from datetime import date
import json
import os
import requests
import click
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Lista wymaganych bibliotek
REQUIRED_PACKAGES = ["requests", "click", "openpyxl"]

# Funkcja do automatycznej instalacji bibliotek
def install_packages():
    for package in REQUIRED_PACKAGES:
        try:
            __import__(package)
        except ImportError:
            print(f"Instalowanie brakującego pakietu: {package}")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Instalacja brakujących bibliotek
install_packages()

# URL API Świąt
HOLIDAYS_API_URL = "https://date.nager.at/api/v2/PublicHolidays/{year}/PL"

def fetch_holidays(year):
    """
    Pobiera listę świąt z API Nager.Date dla podanego roku.
    """
    response = requests.get(HOLIDAYS_API_URL.format(year=year))
    if response.status_code == 200:
        holidays = response.json()
        return {date.fromisoformat(holiday["date"]): holiday["localName"] for holiday in holidays}
    else:
        print(f"Nie udało się pobrać świąt z API. Status: {response.status_code}")
        return {}

def generate_excel_template(year, month, holidays, personal_data, output_path):
    """
    Generuje szablon Excel dla danego miesiąca i roku.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month]}"

    # Nagłówki z danymi osobowymi
    personal_details = [
        ("Dane osoby wystawiającej fakturę:", ""),
        ("Imię i nazwisko:", personal_data.get("name", "")),
        ("Rola/Stanowisko:", personal_data.get("role", "")),
        ("Numer zamówienia/umowy T&M:", personal_data.get("contract", "")),
        ("Stawka godzinowa (PLN):", personal_data.get("hourly_rate", ""))
    ]

    # Zastosowanie scalania komórek dla nagłówków
    ws.merge_cells('A1:B1')  # scalanie dla "Dane osoby wystawiającej fakturę"
    for row_idx, (label, value) in enumerate(personal_details, start=2):
        ws[f"A{row_idx}"] = label
        ws[f"B{row_idx}"] = value
        ws[f"A{row_idx}"].font = Font(bold=True, color="FFFFFF")
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="left")

        # Dodanie obramowania
        ws[f"A{row_idx}"].border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        ws[f"B{row_idx}"].border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    # Zmiana tła na bardziej stonowane kolory
    header_fill = PatternFill(start_color="6FA3EF", end_color="6FA3EF", fill_type="solid")
    for row_idx in range(1, len(personal_details) + 2):
        ws[f"A{row_idx}"].fill = header_fill
        ws[f"B{row_idx}"].fill = header_fill

    start_row = len(personal_details) + 2

    # Nagłówki tabeli
    headers = ["Data", "Liczba godzin", "Opis czynności"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    # Stylizacja komórek
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    weekend_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
    holiday_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
    workday_fill1 = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    workday_fill2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Wypełnienie dni miesiąca
    start_date_row = start_row + 1
    num_days = calendar.monthrange(year, month)[1]
    for i in range(num_days):
        current_date = date(year, month, i + 1)
        row_idx = start_date_row + i
        ws.cell(row=row_idx, column=1, value=current_date.strftime("%Y-%m-%d"))

        # Automatyczne opisy w kolumnie "Opis czynności"
        description_cell = ws.cell(row=row_idx, column=3)
        if current_date in holidays:
            holiday_name = holidays[current_date]  # Pobieranie nazwy święta
            description_cell.value = holiday_name
            fill = holiday_fill
        elif current_date.weekday() == 5:  # Sobota
            description_cell.value = "Sobota"
            fill = weekend_fill
        elif current_date.weekday() == 6:  # Niedziela
            description_cell.value = "Niedziela"
            fill = weekend_fill
        else:
            fill = workday_fill1 if i % 2 == 0 else workday_fill2

        # Zastosowanie obramowania oraz koloru tła dla komórek
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = thin_border

    # Podsumowanie na końcu
    summary_row = start_date_row + num_days
    ws[f"A{summary_row}"] = "Suma godzin:"
    ws[f"A{summary_row}"].font = Font(bold=True)
    ws[f"B{summary_row}"] = f"=SUM(B{start_date_row}:B{start_date_row + num_days - 1})"
    ws[f"B{summary_row}"].font = Font(bold=True)

    # Dodanie obliczenia łącznej kwoty
    ws[f"A{summary_row + 1}"] = "Łączna kwota (PLN):"
    ws[f"A{summary_row + 1}"].font = Font(bold=True)

    # Zastosowanie formuły obliczającej łączną kwotę na podstawie sumy godzin i stawki godzinowej
    ws[f"B{summary_row + 1}"] = f"=B{summary_row} * B6"  # B2 zawiera stawkę godzinową

    ws[f"B{summary_row + 1}"].font = Font(bold=True)

    # Ustawienie szerokości kolumn
    for col_idx in range(1, 4):
        ws.column_dimensions[chr(64 + col_idx)].width = 20

    # Debugowanie - dodanie komunikatu przed zapisem
    print(f"Zapisuję plik pod ścieżką: {output_path}")
    
    # Zapis pliku
    try:
        wb.save(output_path)
        print(f"Plik zapisany pomyślnie.")
    except Exception as e:
        print(f"Wystąpił błąd podczas zapisywania pliku: {e}")

def load_config():
    if os.path.exists("person_config.json"):
        with open("person_config.json", "r") as f:
            return json.load(f)
    return {}

def save_config(data):
    with open("person_config.json", "w") as f:
        json.dump(data, f, indent=4)

@click.group()
def cli():
    pass

@cli.command()
@click.option("--name", prompt="Imię i nazwisko", help="Imię i nazwisko.")
@click.option("--role", prompt="Rola/Stanowisko", help="Rola osoby.")
@click.option("--contract", prompt="Numer zamówienia/umowy T&M", help="Numer zamówienia.")
@click.option("--hourly_rate", prompt="Stawka godzinowa (PLN)", type=float, help="Stawka godzinowa.")
def configure(name, role, contract, hourly_rate):
    config = {"name": name, "role": role, "contract": contract, "hourly_rate": hourly_rate}
    save_config(config)
    click.echo("Dane zapisane.")

@cli.command()
@click.option("--year", prompt="Rok", type=int, help="Rok raportu.")
@click.option("--month", prompt="Miesiąc", type=int, help="Miesiąc raportu (1-12).")
def generate(year, month):
    config = load_config()
    if not config:
        click.echo("Brak danych osoby. Uruchom 'configure'.")
        return

    holidays = fetch_holidays(year)
    output_path = f"raport_{calendar.month_name[month]}_{year}.xlsx"
    generate_excel_template(year, month, holidays, config, output_path)

if __name__ == "__main__":
    cli()
